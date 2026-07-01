"""
exporter.py
===========
Writes the generated corpus to all required output formats:
  * requirements.json          -> list of plain strings (Ollama input;
                                   ONLY the generated_requirement field)
  * requirements.csv           -> full structured records
  * requirements.xlsx          -> full structured records, formatted
  * requirements_metadata.json -> full structured records as JSON
  * generation_statistics.json -> run statistics for the IEEE paper
"""

from __future__ import annotations

import csv
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import GeneratorConfig
from metadata import FirewallRequirement
from generator import GenerationStatistics


def _ensure_output_dir(config: GeneratorConfig) -> None:
    paths = config.output_paths()
    out_dir = os.path.dirname(next(iter(paths.values())))
    os.makedirs(out_dir, exist_ok=True)


def export_requirements_json(requirements: list[FirewallRequirement], path: str) -> None:
    """Exports ONLY the generated_requirement strings, in the format
    expected as direct LLM input for the Ollama / Llama 3.1 stage."""
    payload = [r.generated_requirement for r in requirements]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


def export_metadata_json(requirements: list[FirewallRequirement], path: str) -> None:
    payload = [r.to_dict() for r in requirements]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


_CSV_FIELDS = [
    "id", "category", "subcategory", "difficulty", "template_id",
    "generated_requirement", "service", "protocol", "port",
    "source_zone", "destination_zone", "source_host", "destination_host",
    "source_cidr", "destination_cidr", "department", "application",
    "vendor", "compliance", "authentication", "action", "internet_exposed",
    "num_conditions", "num_exceptions", "has_nested_logic",
    "has_contradiction", "has_ambiguity", "ambiguity_score",
    "complexity_score",
]


def export_requirements_csv(requirements: list[FirewallRequirement], path: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for r in requirements:
            row = r.to_dict()
            writer.writerow({k: row.get(k) for k in _CSV_FIELDS})


def export_requirements_xlsx(requirements: list[FirewallRequirement], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for col_idx, field_name in enumerate(_CSV_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(requirements, start=2):
        row = r.to_dict()
        for col_idx, field_name in enumerate(_CSV_FIELDS, start=1):
            value = row.get(field_name)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Reasonable column widths
    widths = {
        "id": 12, "category": 18, "subcategory": 18, "difficulty": 10,
        "template_id": 18, "generated_requirement": 90, "service": 28,
        "protocol": 10, "port": 8, "source_zone": 26, "destination_zone": 26,
        "source_host": 16, "destination_host": 16, "source_cidr": 16,
        "destination_cidr": 16, "department": 22, "application": 26,
        "vendor": 18, "compliance": 14, "authentication": 30, "action": 10,
        "internet_exposed": 12, "num_conditions": 10, "num_exceptions": 10,
        "has_nested_logic": 10, "has_contradiction": 12, "has_ambiguity": 10,
        "ambiguity_score": 10, "complexity_score": 10,
    }
    for col_idx, field_name in enumerate(_CSV_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(field_name, 14)

    ws.freeze_panes = "A2"
    wb.save(path)


def export_statistics_json(stats: GenerationStatistics, path: str) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(stats.to_dict(), fh, indent=2, ensure_ascii=False)


def export_all(
    requirements: list[FirewallRequirement],
    stats: GenerationStatistics,
    config: GeneratorConfig,
) -> dict:
    _ensure_output_dir(config)
    paths = config.output_paths()

    export_requirements_json(requirements, paths["requirements_json"])
    export_metadata_json(requirements, paths["metadata_json"])
    export_requirements_csv(requirements, paths["requirements_csv"])
    export_requirements_xlsx(requirements, paths["requirements_xlsx"])
    export_statistics_json(stats, paths["statistics_json"])

    return paths
